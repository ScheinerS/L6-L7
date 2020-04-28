#%%
# IMPORTO LIBRERÍAS PARA QUE FUNCIONE TODO. DESPUÉS HAY QUE SACARLAS:

import os
import sys
import pandas as pd
import openpyxl
import numpy as np
from datetime import timedelta


'''
Para el path absoluto, correr el programa con F5, y no bloque por bloque.
'''

# path0 = '/Users/Tele/Desktop/L6-L7/FLNTU/Base de datos'

# path0      = '/home/gossn/Dropbox/Documents/L6y7_Scheiner_Santamaria/Datos'
# pathModules = '/home/gossn/Dropbox/Documents/L6y7_Scheiner_Santamaria/scripts'

path0 = os.path.dirname(os.path.realpath('__file__'))
pathModules = os.path.dirname(path0) + '/scripts'
path0 = os.path.dirname(path0) + '/Datos'

sys.path.append(pathModules)

import ECO_DataCleaner as DC

campaign0 = 'RdP_20191217_Muelle'

#%%
def ECO2Stations(campaign0,path0):

    # campaign0 = 'RdP_20191217_Muelle'
    # path0      = '/home/gossn/Dropbox/Documents/L6y7_Scheiner_Santamaria/Datos'

# en el siguiente renglon el path no es a los modulos, sino a los datos:
    # pathRegions   = pathModules + '/regions'
    pathRegions   = path0 + '/regions'

    region = campaign0.split('_')[0]
    date   = campaign0.split('_')[1]
    campaign = region + '_' + date
    pathCampaign  = pathRegions + '/' + region + '/' +  campaign0

    
    print('Processing ECO data for campaign: ' + campaign0)
    
    inputs = {}
    try:
        with open(pathCampaign + '/ECO_FLNTU/ECO_FLNTUProcessingInputs') as f0:
            for line in f0:
               (key, val) = line.split()
               inputs[key] = val.split(',')
               if len(inputs[key])==1:
                   inputs[key] = inputs[key][0]
        
        # Limpiamos el archivo del ECO:
        DC.clean(pathCampaign)
    except:
        print('No ECO Input file or no ECO measurements for this campaign!')
#        return
    if not os.path.isdir(pathCampaign + '/ECO_FLNTUProcessed/'):
        os.mkdir(pathCampaign + '/ECO_FLNTUProcessed/')

    # ECO Processing
    
    if not os.path.isdir(pathCampaign + '/ECO_FLNTU'):
        print("No ECO data available for this campaign!")
    else:
        filenameCs = campaign + '_cleaned.xlsx'
        
        # Read station IDs and Times
        stationInfo = pd.DataFrame()
        stationInfo = pd.read_excel(pathCampaign + '/' + campaign + '.xlsx',sheet_name='stationInfo',skiprows=1)

        stationIDs   = stationInfo['StationID'   ].astype(object)
        stationTimes = stationInfo['timeStampUTC'].astype(object)
    
    
        pathXlsx = pathCampaign + '/ECO_FLNTUProcessed/' + campaign + '_ECO-FLNTU.xlsx'
        wb = openpyxl.Workbook()
        wb.save(pathXlsx)
        writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
        writer.book = wb
        del wb['Sheet']
    
        # Write to Excel: StationInfo
        sheetname = 'stationInfo'
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        stationInfo.set_index('StationID')
        stationInfo.to_excel(writer, index = False, sheet_name=sheetname,startrow = 1)

        # Initialize dictionary of dataFrame that will store data per station (all dataloggers)
        csStations = {}
        for stat in ['Mean','Std','CV']:
            csStations[stat] = pd.DataFrame(index=stationIDs)

        csCont = pd.DataFrame()
        
        csCont = pd.read_excel(pathCampaign + '/ECO_FLNTU/' + filenameCs, header = 0,index_col = 0)
        
        '''
        # Change column names
        colNamesOld = list(csCont.columns)
        colNamesNew = []
        for cname in colNamesOld:
            if cname[1][:len('Unnamed')] == 'Unnamed':
                colNamesNew.append(cname[0])
            else:
                colNamesNew.append(cname[0] + '[' + cname[1].replace(' ','') + ']')
        csCont.columns = colNamesNew
        '''

        # ECO Times
        # Corrección por desfasaje entre huso horario del instrumento y UTC:
        csCont['timestamp'] = pd.to_datetime(csCont['timestamp'])-timedelta(hours=float(inputs['deltaUTC']))
        csContTime = csCont['timestamp']
        
        ##### DATOS POR ESTACION
        print('Data per station')
        
        # ECO collected data

        # csContData = pd.DataFrame()

        # others = [c for c in csCont.columns if (c.lower()[:2] == 'wd' or c.lower()[:12] == 'stationnames' or c.lower()[:7] == 'temp_cr')]

        # J: En 'csContData' van las variables a promediar, hay que sacar los tiempos y los parametros fijos
        csContData = csCont.copy()
        csContData = csCont.drop(['date','time','cpu_temperature', 'wavelength_turbidity', 'wavelength_chl_emission','timestamp', 'turbidity_counts', 'chl_counts'], axis=1)
        csContData = csContData.apply(pd.to_numeric, errors='coerce')
        
        csMeasures = list(csContData.columns.values)

        csStStats  = {}
        st0 = -1
        for st in stationIDs:
            st0+=1
            print('Processing station ' + str(st))
            timeSt = stationTimes[stationIDs==st][st0]
            timeDeltaSt = abs(csContTime-timeSt)<pd.Timedelta(float(inputs['timeDeltaStationMin']), unit='m')
            if any(timeDeltaSt):
                csContSt = csContData.loc[timeDeltaSt,:]
                Q1 = csContSt.quantile(0.25)
                Q3 = csContSt.quantile(0.75)
                IQR = Q3 - Q1
                outliers = (csContSt < (Q1 - 1.5 * IQR)) |(csContSt > (Q3 + 1.5 * IQR))
                csContSt[outliers] = np.nan
                csStMed = csContSt.mean()
                csStStd = csContSt.std()
                csStCV  = csStStd/csStMed*100
                
                csStStats[st] = {'Mean': csStMed, 'Std': csStStd,'CV': csStCV}
                
        for stat in ['Mean','Std','CV']:
            csStationsFile = pd.DataFrame(columns=csMeasures)
            for st in stationIDs:
                try:
                    csStationsFile.loc[st]   = csStStats[st][stat]
                except:
                    csStationsFile.loc[st]   = pd.Series(index=csMeasures)
            csStationsFile.fillna('')
            csStationsFile.index.name = 'StationID'
            csStations[stat] = pd.concat([csStations[stat], csStationsFile], axis=1)
                                

            # Por ahora olvídense de este bloque
            # ##### DATOS CRUDOS
            # # J: El siguiente renglon claramente hay q cambiarlo:
            # # sheetname = file[:-4] # comentado por J
            # sheetname = 'ECO_FLNTU_continuous' + '_' + stat
            # if sheetname in wb.sheetnames:
            #     del wb[sheetname]
            # csContData.to_excel(writer, index = False, sheet_name=sheetname)
            # #adjustColWidth(wb.get_sheet_by_name(sheetname))
            # writer.save()
            # writer.close()
            

        for stat in ['Mean','Std','CV']:
            csStations[stat] = csStations[stat].reindex(columns=sorted(csStations[stat].columns))
            sheetname = 'Stations' + '_' + stat
            if sheetname in wb.sheetnames:
                del wb[sheetname]
            csStations[stat].to_excel(writer, index = True, sheet_name=sheetname, startrow=1)
            #adjustColWidth(wb.get_sheet_by_name(sheetname))
            writer.save()
            writer.close()

        ##### DATOS SUAVIZADOS A 'smoothWinMin' MIN
    #if file[0:6] != 'CR200X': # ESTE IF ELIMINENLO! ES el q les eliminé sin querer.
        csSmooth = pd.DataFrame()
        step = 0
        flag = True
        while flag:
            
            timeWin = csCont['timestamp'].min() + timedelta(minutes=step*float(inputs['smoothWinMin']))
            step+=1
            flag = timeWin < csCont['timestamp'].max()
            timeDeltaWin = abs(csContTime-timeWin)<pd.Timedelta(float(inputs['smoothWinMin'])/2, unit='m')
            if any(timeDeltaWin):
                csContWin = csContData.loc[timeDeltaWin,:]
                Q1 = csContWin.quantile(0.25)
                Q3 = csContWin.quantile(0.75)
                IQR = Q3 - Q1
                outliers = (csContWin < (Q1 - 1.5 * IQR)) |(csContWin > (Q3 + 1.5 * IQR))
                csContWin[outliers] = np.nan
                csWinMedia = csContWin.mean()
                csWinStd = csContWin.std()
                csWinCV = csWinStd/csWinMedia*100
                
                #csWinStats = [csWinMedia,csWinStd,csWinMedia]
                
                csMeasures = list(csContWin.columns.values)
                for m in csMeasures:
                    csSmooth.loc[timeWin,m + 'Mean'  ] = csWinMedia[m]
                    csSmooth.loc[timeWin,m + 'Std'  ] = csWinStd[m]
                    csSmooth.loc[timeWin,m + 'CV'] = csWinCV[m]
        # ¿Qué es 'file'? ¿El nombre del archivo? La línea no anda porque nos falta 'file':
        sheetname = 'ECO' + 'ContSmooth' + inputs['smoothWinMin'] + 'min'
        
        if sheetname in wb.sheetnames:
            del wb[sheetname]
        csSmooth.to_excel(writer, index = True, sheet_name=sheetname)
        #adjustColWidth(wb[sheetname])
        writer.save()
        writer.close()

#%%

# CORRO LA FUNCIÓN. BORRAR ESTE BLOQUE CUANDO ESTÉ TERMINADO.

ECO2Stations(campaign0,path0)

