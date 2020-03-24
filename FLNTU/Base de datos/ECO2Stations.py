# IMPORTO LIBRERÍAS PARA QUE FUNCIONE TODO. DESPUÉS HAY QUE SACARLAS:

import os
import sys
import pandas as pd
import openpyxl
import numpy as np
from datetime import timedelta


# path0 = '/Users/Tele/Desktop/L6-L7/FLNTU/Base de datos'

path0 = os.path.dirname(os.path.realpath('__file__'))
sys.path.append(path0)

campaign0 = 'RdP_20191217_Muelle'


#%%

# Eliminamos las líneas con errores:
# pongo esto acá porque necesito el path0. Necesito pasarle el path0 al archivo y no entiendo cómo...

import ECO_Corrector.py as corr

# Hay armar una función que corra todo dentro de ECO_Corrector y va a haber que pasarle los directorios como string.

def ECO2Stations(campaign0,path0):
    campaign0 = 'RdP_20191217_Muelle' 

    pathRegions   = path0 + '/regions'

    region = campaign0.split('_')[0]
    date   = campaign0.split('_')[1]
    campaign = region + '_' + date
    pathCampaign  = pathRegions + '/' + region + '/' +  campaign0
    
    print('Processing ECO data for campaign: ' + campaign0)
    
    inputs = {}
    try:
        with open(pathCampaign + '/ECO_FLNTU/ECO_FLNTUProcessingInputs') as f0:
            for line in f0:
               (key, val) = line.split()
               inputs[key] = val.split(',')
               if len(inputs[key])==1:
                   inputs[key] = inputs[key][0]
    except:
        print('No ECO Input file or no ECO measurements for this campaign!')
        #return
    if not os.path.isdir(pathCampaign + '/ECO_FLNTUProcessed/'):
        os.mkdir(pathCampaign + '/ECO_FLNTUProcessed/')
    
    # OBS Processing
    
    if not os.path.isdir(pathCampaign + '/ECO_FLNTU'):
        print("No ECO data available for this campaign!")
    else:
        # Append sheet to Excel logsheet with Campbell Sci. continuous measurements...
        filenameCs = os.listdir(pathCampaign + '/ECO_FLNTU')
        filenameCs.remove('ECO_FLNTUProcessingInputs')
        
        # Read station IDs and Times
        stationInfo = pd.DataFrame()
        stationInfo = pd.read_excel(pathCampaign + '/' + campaign + '.xlsx',sheet_name='stationInfo',skiprows=1)
    
        # Remendar bug openpyxl (borra formatos preestablecidos)
        stationInfo['startTimeUTC'] = pd.to_datetime(stationInfo['startTimeUTC'],format= '%H:%M:%S' )
        stationInfo['timeStampUTC'] = stationInfo['DateUTC'] + pd.to_timedelta(stationInfo['startTimeUTC']) + timedelta(hours=70*24*365.25-12)
        
        stationIDs   = stationInfo['StationID'].asobject
        stationTimes = stationInfo['timeStampUTC'].asobject
    
    
        pathXlsx = pathCampaign + '/ECO_FLNTUProcessed/' + campaign + '_ECO-FLNTU.xlsx'
        wb = openpyxl.Workbook()
        wb.save(pathXlsx)
        writer = pd.ExcelWriter(pathXlsx, engine = 'openpyxl')
        writer.book = wb
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))   
    
        # Write to Excel: StationInfo
        sheetname = 'stationInfo'
        if sheetname in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        stationInfo.set_index('StationID')
        stationInfo.to_excel(writer, index = False, sheet_name=sheetname,startrow = 1)

        # Initialize dictionary of dataFrame that will store data per station (all dataloggers)
        csStations = {}
        for stat in ['Mean','Std','CV']:
            csStations[stat] = pd.DataFrame(index=stationIDs)
    
        for file in filenameCs:
            
            csCont = pd.DataFrame()
            csCont = pd.read_csv(pathCampaign + '/ECO_FLNTU' + '/' + file, header = [1,2])
    
            # Change column names
            colNamesOld = list(csCont.columns)
            colNamesNew = []
            for cname in colNamesOld:
                if cname[1][:len('Unnamed')] == 'Unnamed':
                    colNamesNew.append(cname[0])
                else:
                    colNamesNew.append(cname[0] + '[' + cname[1].replace(' ','') + ']')
            csCont.columns = colNamesNew
    
            # campbell Times
            csCont = csCont.drop(csCont.index[[0]])    # Drop extra-headers
            csCont['TIMESTAMP[TS]'] = pd.to_datetime(csCont['TIMESTAMP[TS]'])-timedelta(hours=float(inputs['deltaUTC']))
            csContTime = csCont['TIMESTAMP[TS]']
    
            
            ##### DATOS POR ESTACION
            print('Data per station')
            
            # campbell collected data
            csContData = pd.DataFrame()
            others = [c for c in csCont.columns if (c.lower()[:2] == 'wd' or c.lower()[:12] == 'stationnames' or c.lower()[:7] == 'temp_cr')]
            csContData = csCont.drop(['TIMESTAMP[TS]','RECORD[RN]','BattV[Volts]'] + others, axis=1)
            csContData = csContData.apply(pd.to_numeric, errors='coerce')
            csMeasures = list(csContData.columns.values)
        
            csStStats  = {}
            for st in stationIDs:
                print('Processing station ' + str(st))
                timeSt = stationTimes[stationIDs==st][0]
                timeDeltaSt = abs(csContTime-timeSt)<pd.Timedelta(float(inputs['timeDeltaStationMin']), unit='m')
                if any(timeDeltaSt):
                    csContSt = csContData.loc[timeDeltaSt,:]
                    Q1 = csContSt.quantile(0.25)
                    Q3 = csContSt.quantile(0.75)
                    IQR = Q3 - Q1
                    outliers = (csContSt < (Q1 - 1.5 * IQR)) |(csContSt > (Q3 + 1.5 * IQR))
                    csContSt[outliers] = np.nan
                    csStMed = csContSt.mean()
                    csStStd = csContSt.std()
                    csStCV  = csStStd/csStMed*100
                    
                    csStStats[st] = {'Mean': csStMed, 'Std': csStStd,'CV': csStCV}
                    
            for stat in ['Mean','Std','CV']:
                csStationsFile = pd.DataFrame(columns=csMeasures)
                for st in stationIDs:
                    try:
                        csStationsFile.loc[st]   = csStStats[st][stat]
                    except:
                        csStationsFile.loc[st]   = pd.Series(index=csMeasures)
                csStationsFile.fillna('')
                csStationsFile.index.name = 'StationID'
                csStations[stat] = pd.concat([csStations[stat], csStationsFile], axis=1)
                                

            ##### DATOS CRUDOS
            sheetname = file[:-4]
            if sheetname in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
            csCont.to_excel(writer, index = False, sheet_name=sheetname)
            #adjustColWidth(wb.get_sheet_by_name(sheetname))
            writer.save()
            writer.close()
            
            ##### DATOS SUAVIZADOS A 'smoothWinMin' MIN
            if file[0:6] != 'CR200X':
                csSmooth = pd.DataFrame()
                step = 0
                flag = True
                while flag:
                    timeWin = csCont['TIMESTAMP[TS]'].min() + timedelta(minutes=step*float(inputs['smoothWinMin']))
                    step+=1
                    flag = timeWin < csCont['TIMESTAMP[TS]'].max()
                    timeDeltaWin = abs(csContTime-timeWin)<pd.Timedelta(float(inputs['smoothWinMin'])/2, unit='m')
                    if any(timeDeltaWin):
                        csContWin = csContData.loc[timeDeltaWin,:]
                        Q1 = csContWin.quantile(0.25)
                        Q3 = csContWin.quantile(0.75)
                        IQR = Q3 - Q1
                        outliers = (csContWin < (Q1 - 1.5 * IQR)) |(csContWin > (Q3 + 1.5 * IQR))
                        csContWin[outliers] = np.nan
                        csWinMedia = csContWin.mean()
                        csWinStd = csContWin.std()
                        csWinCV = csWinStd/csWinMedia*100
                        
#                        csWinStats = [csWinMedia,csWinStd,csWinMedia]
                        
                        csMeasures = list(csContWin.columns.values)
                        for m in csMeasures:
                            csSmooth.loc[timeWin,m + 'Mean'  ] = csWinMedia[m]
                            csSmooth.loc[timeWin,m + 'Std'  ] = csWinStd[m]
                            csSmooth.loc[timeWin,m + 'CV'] = csWinCV[m]
                sheetname = file[:-4] + 'ContSmooth' + inputs['smoothWinMin'] + 'min'
                if sheetname in wb.sheetnames:
                    wb.remove_sheet(wb.get_sheet_by_name(sheetname))
                csSmooth.to_excel(writer, index = True, sheet_name=sheetname)
                #adjustColWidth(wb.get_sheet_by_name(sheetname))
                writer.save()
                writer.close()


        for stat in ['Mean','Std','CV']:
            csStations[stat] = csStations[stat].reindex(columns=sorted(csStations[stat].columns))
            sheetname = 'Stations' + '_' + stat
            if sheetname in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sheetname))
            csStations[stat].to_excel(writer, index = True, sheet_name=sheetname, startrow=1)
            #adjustColWidth(wb.get_sheet_by_name(sheetname))
            writer.save()
            writer.close()

#%%

# CORRO LA FUNCIÓN. BORRAR ESTE BLOQUE CUANDO ESTÉ TERMINADO.

ECO2Stations(campaign0,path0)